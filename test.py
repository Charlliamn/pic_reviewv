from __future__ import annotations

import argparse
import base64
import json
import os
import re
import sys
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

import openpyxl
import requests


DEFAULT_MODEL = "gpt-4o-mini"
SYSTEM_PROMPT = (
    "你是一名严谨的摄影图像分析助手。只根据图片本身分析，不要编造不可见信息。"
    "输出必须是 JSON 对象。"
)
USER_PROMPT = """请分析这张图片，返回 JSON：
{
  "light_shadow": "光影分析，包含光源方向、明暗关系、影调、氛围",
  "main_idea": "图片主旨/可能表达的核心",
  "objects": ["画面中主要对象，按重要性排序"],
  "composition": "构图与视线组织的简短分析",
  "confidence": "high/medium/low"
}
要求使用中文，简洁但具体。"""


@dataclass
class ExtractedImage:
    sheet: str
    row: int
    record_no: str
    time_value: str
    location: str
    image_index: int
    anchor_col: int
    file_path: str
    mime_type: str
    width: int | None
    height: int | None


def find_workbook(workbook_arg: str | None) -> Path:
    if workbook_arg:
        path = Path(workbook_arg)
        if not path.exists():
            raise FileNotFoundError(f"找不到指定的 xlsx: {path}")
        return path

    candidates = [p for p in Path(".").glob("*.xlsx") if not p.name.startswith("~$")]
    if not candidates:
        raise FileNotFoundError("当前目录没有找到 .xlsx 文件")
    return max(candidates, key=lambda p: p.stat().st_mtime)


def load_env_file(path: Path = Path(".env")) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def safe_name(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r'[<>:"/\\|?*\s]+', "_", text)
    text = text.strip("._")
    return text[:80] or "unknown"


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def get_header_map(ws: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[1]:
        value = cell_text(cell.value)
        if value:
            headers[value] = cell.column
    return headers


def get_picture_columns(ws: Any, image_cols: set[int]) -> set[int]:
    headers = get_header_map(ws)
    picture_col = headers.get("图片")
    if not picture_col:
        return image_cols

    allowed = {picture_col}
    col = picture_col + 1
    while col <= ws.max_column:
        header_value = cell_text(ws.cell(row=1, column=col).value)
        if header_value:
            break
        if col in image_cols:
            allowed.add(col)
        col += 1
    return allowed


def image_bytes(img: Any) -> bytes:
    data = img._data()
    if not data:
        raise ValueError("图片数据为空")
    return data


def image_extension_and_mime(img: Any, data: bytes) -> tuple[str, str]:
    fmt = (getattr(img, "format", None) or "").lower()
    if fmt in {"jpg", "jpeg"}:
        return "jpg", "image/jpeg"
    if fmt == "png":
        return "png", "image/png"
    if data.startswith(b"\xff\xd8"):
        return "jpg", "image/jpeg"
    if data.startswith(b"\x89PNG\r\n\x1a\n"):
        return "png", "image/png"
    return fmt or "bin", "application/octet-stream"


def extract_images(workbook: Path, output_dir: Path) -> list[ExtractedImage]:
    wb = openpyxl.load_workbook(workbook, data_only=True)
    extracted: list[ExtractedImage] = []

    for ws in wb.worksheets:
        images = getattr(ws, "_images", [])
        anchored = []
        for original_order, img in enumerate(images, start=1):
            anchor = img.anchor
            if not hasattr(anchor, "_from"):
                continue
            anchored.append(
                {
                    "img": img,
                    "original_order": original_order,
                    "row": anchor._from.row + 1,
                    "col": anchor._from.col + 1,
                }
            )

        if not anchored:
            continue

        image_cols = {item["col"] for item in anchored}
        picture_cols = get_picture_columns(ws, image_cols)
        headers = get_header_map(ws)
        no_col = headers.get("序号", 1)
        time_col = headers.get("时间")
        location_col = headers.get("地点")

        by_row: dict[int, list[dict[str, Any]]] = {}
        for item in anchored:
            if item["row"] <= 1 or item["col"] not in picture_cols:
                continue
            by_row.setdefault(item["row"], []).append(item)

        for row, row_images in sorted(by_row.items()):
            record_no = cell_text(ws.cell(row=row, column=no_col).value) or str(row - 1)
            time_value = cell_text(ws.cell(row=row, column=time_col).value) if time_col else ""
            location = cell_text(ws.cell(row=row, column=location_col).value) if location_col else ""
            row_images.sort(key=lambda item: (item["col"], item["original_order"]))

            for image_index, item in enumerate(row_images, start=1):
                img = item["img"]
                data = image_bytes(img)
                ext, mime_type = image_extension_and_mime(img, data)
                sheet_dir = output_dir / safe_name(ws.title)
                image_dir = sheet_dir / "images"
                image_dir.mkdir(parents=True, exist_ok=True)
                filename = (
                    f"{safe_name(ws.title)}_row{row:04d}_no{safe_name(record_no)}"
                    f"_image{image_index}.{ext}"
                )
                file_path = image_dir / filename
                file_path.write_bytes(data)

                extracted.append(
                    ExtractedImage(
                        sheet=ws.title,
                        row=row,
                        record_no=record_no,
                        time_value=time_value,
                        location=location,
                        image_index=image_index,
                        anchor_col=item["col"],
                        file_path=str(file_path),
                        mime_type=mime_type,
                        width=getattr(img, "width", None),
                        height=getattr(img, "height", None),
                    )
                )

    return extracted


def group_images_by_sheet(images: list[ExtractedImage]) -> dict[str, list[ExtractedImage]]:
    groups: dict[str, list[ExtractedImage]] = {}
    for image in images:
        groups.setdefault(image.sheet, []).append(image)
    return groups


def parse_json_response(text: str) -> dict[str, Any]:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    try:
        value = json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, flags=re.S)
        if not match:
            raise
        value = json.loads(match.group(0))
    if not isinstance(value, dict):
        raise ValueError("API 返回不是 JSON 对象")
    return value


def image_result_key(image: ExtractedImage) -> str:
    return f"{image.sheet}|{image.row}|{image.record_no}|{image.image_index}"


def row_result_key(row: dict[str, Any]) -> str:
    return f"{row.get('sheet')}|{row.get('row')}|{row.get('record_no')}|{row.get('image_index')}"


def load_existing_results(jsonl_path: Path) -> dict[str, dict[str, Any]]:
    if not jsonl_path.exists():
        return {}

    results: dict[str, dict[str, Any]] = {}
    with jsonl_path.open("r", encoding="utf-8") as fp:
        for line in fp:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except json.JSONDecodeError:
                continue
            if row.get("analysis") and not row.get("error"):
                results[row_result_key(row)] = row
    return results


def retry_delay_seconds(exc: Exception, attempt: int) -> float:
    if isinstance(exc, requests.HTTPError) and exc.response is not None:
        retry_after = exc.response.headers.get("Retry-After")
        if retry_after:
            try:
                return max(float(retry_after), 1.0)
            except ValueError:
                pass
        if exc.response.status_code == 429:
            return min(60.0, 5.0 * (2 ** (attempt - 1)))
        if 500 <= exc.response.status_code < 600:
            return min(30.0, 3.0 * (2 ** (attempt - 1)))
    return min(20.0, 2.0 * attempt)


def call_vision_api(image: ExtractedImage, api_key: str, model: str, base_url: str) -> dict[str, Any]:
    data = Path(image.file_path).read_bytes()
    encoded = base64.b64encode(data).decode("ascii")
    payload: dict[str, Any] = {
        "model": model,
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": USER_PROMPT},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:{image.mime_type};base64,{encoded}"},
                    },
                ],
            },
        ],
        "temperature": 0.2,
        "response_format": {"type": "json_object"},
    }

    url = f"{base_url.rstrip('/')}/chat/completions"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    response = requests.post(url, headers=headers, json=payload, timeout=120)
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        if response.status_code in {400, 422} and "response_format" in response.text:
            payload.pop("response_format", None)
            response = requests.post(url, headers=headers, json=payload, timeout=120)
            response.raise_for_status()
        else:
            raise exc

    content = response.json()["choices"][0]["message"]["content"]
    return parse_json_response(content)


def analyze_images(
    images: list[ExtractedImage],
    output_dir: Path,
    model: str,
    base_url: str,
    limit: int | None,
    sleep_seconds: float,
    max_retries: int,
    resume: bool,
) -> list[dict[str, Any]]:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("缺少 OPENAI_API_KEY。请先配置环境变量，或使用 --extract-only 只抽取图片。")
    if "api.deepseek.com" in base_url.lower():
        raise RuntimeError(
            "当前任务需要视觉模型分析图片，但 DeepSeek 官方 API 目前不是图片视觉接口。"
            "请改用支持 image_url 的视觉模型接口，例如 OpenAI gpt-4o-mini/gpt-4o，"
            "或其他 OpenAI 兼容的多模态模型服务。"
        )

    targets = images[:limit] if limit else images
    jsonl_path = output_dir / "analysis_results.jsonl"
    existing = load_existing_results(jsonl_path) if resume else {}
    results: list[dict[str, Any]] = list(existing.values())
    mode = "a" if resume else "w"

    with jsonl_path.open(mode, encoding="utf-8") as fp:
        for index, image in enumerate(targets, start=1):
            key = image_result_key(image)
            if key in existing:
                print(f"[{index}/{len(targets)}] 已存在，跳过 {image.sheet} 序号={image.record_no} 第{image.image_index}张")
                continue

            print(f"[{index}/{len(targets)}] 分析 {image.sheet} 序号={image.record_no} 第{image.image_index}张")
            last_error = None
            analysis = None
            for attempt in range(1, max_retries + 1):
                try:
                    analysis = call_vision_api(image, api_key=api_key, model=model, base_url=base_url)
                    break
                except Exception as exc:
                    last_error = exc
                    if attempt == max_retries:
                        break
                    delay = retry_delay_seconds(exc, attempt)
                    print(f"  请求失败：{exc}；{delay:.0f} 秒后重试 ({attempt}/{max_retries})")
                    time.sleep(delay)

            row = asdict(image)
            row["analysis"] = analysis
            row["error"] = None if analysis is not None else str(last_error)
            results.append(row)
            fp.write(json.dumps(row, ensure_ascii=False) + "\n")
            fp.flush()
            if sleep_seconds:
                time.sleep(sleep_seconds)
    return results


def write_manifest(images: list[ExtractedImage], output_dir: Path) -> Path:
    path = output_dir / "extracted_images_manifest.json"
    rows = [asdict(image) for image in images]
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def write_excel(results: list[dict[str, Any]], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "analysis"
    append_results_to_sheet(ws, results)
    adjust_sheet_widths(ws)

    path = output_dir / "analysis_results.xlsx"
    wb.save(path)
    return path


def append_results_to_sheet(ws: Any, results: list[dict[str, Any]]) -> None:
    headers = [
        "sheet",
        "excel_row",
        "序号",
        "时间",
        "地点",
        "图片序号",
        "图片文件",
        "光影",
        "主旨",
        "画面对象",
        "构图",
        "置信度",
        "错误",
    ]
    ws.append(headers)

    for item in results:
        analysis = item.get("analysis") or {}
        objects = analysis.get("objects", "")
        if isinstance(objects, list):
            objects = "、".join(str(x) for x in objects)
        ws.append(
            [
                item.get("sheet"),
                item.get("row"),
                item.get("record_no"),
                item.get("time_value"),
                item.get("location"),
                item.get("image_index"),
                item.get("file_path"),
                analysis.get("light_shadow", ""),
                analysis.get("main_idea", ""),
                objects,
                analysis.get("composition", ""),
                analysis.get("confidence", ""),
                item.get("error"),
            ]
        )


def adjust_sheet_widths(ws: Any) -> None:
    for col in ws.columns:
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in col) + 2, 60)


def write_combined_excel(results_by_sheet: dict[str, list[dict[str, Any]]], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_titles: set[str] = set()
    source = results_by_sheet or {"analysis": []}
    for sheet_name, results in source.items():
        title = unique_excel_sheet_title(sheet_name, used_titles)
        ws = wb.create_sheet(title=title)
        append_results_to_sheet(ws, results)
        adjust_sheet_widths(ws)

    path = output_dir / "analysis_results_by_document.xlsx"
    wb.save(path)
    return path


def unique_excel_sheet_title(value: Any, used_titles: set[str]) -> str:
    title = str(value or "analysis").strip()
    title = re.sub(r"[\[\]:*?/\\]", "_", title)
    title = title[:31].strip() or "analysis"
    base = title
    index = 2
    while title in used_titles:
        suffix = f"_{index}"
        title = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(title)
    return title


def main() -> int:
    load_env_file()

    parser = argparse.ArgumentParser(description="读取 xlsx 中图片列的图片，并调用视觉 API 分析。")
    parser.add_argument("--workbook", help="xlsx 文件路径；不传则自动使用当前目录最近修改的 .xlsx")
    parser.add_argument("--output-dir", default="output", help="输出目录")
    parser.add_argument("--model", default=os.getenv("OPENAI_MODEL", DEFAULT_MODEL), help="视觉模型名")
    parser.add_argument(
        "--base-url",
        default=os.getenv("OPENAI_BASE_URL", "https://api.openai.com/v1"),
        help="OpenAI 兼容接口 base url",
    )
    parser.add_argument("--extract-only", action="store_true", help="只抽取图片，不调用 API")
    parser.add_argument("--limit", type=int, help="只分析前 N 张图片，方便测试")
    parser.add_argument("--sleep", type=float, default=0.0, help="每张图分析后的等待秒数")
    parser.add_argument("--max-retries", type=int, default=6, help="每张图片失败后的最大重试次数")
    parser.add_argument("--no-resume", action="store_true", help="不读取已有 analysis_results.jsonl，强制重新分析")
    args = parser.parse_args()

    workbook = find_workbook(args.workbook)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"使用 xlsx：{workbook}")
    images = extract_images(workbook, output_dir)
    images_by_sheet = group_images_by_sheet(images)
    print(f"已抽取 {len(images)} 张图片，分文档输出到：{output_dir}")
    for sheet, sheet_images in images_by_sheet.items():
        sheet_output_dir = output_dir / safe_name(sheet)
        manifest_path = write_manifest(sheet_images, sheet_output_dir)
        print(f"  {sheet}: {len(sheet_images)} 张图片，清单：{manifest_path}")

    if args.extract_only:
        return 0

    remaining_limit = args.limit
    results_by_sheet: dict[str, list[dict[str, Any]]] = {}
    for sheet, sheet_images in images_by_sheet.items():
        if remaining_limit is not None and remaining_limit <= 0:
            break

        sheet_limit = None
        if remaining_limit is not None:
            sheet_limit = min(remaining_limit, len(sheet_images))
            remaining_limit -= sheet_limit

        sheet_output_dir = output_dir / safe_name(sheet)
        print(f"开始分析文档：{sheet}")
        results = analyze_images(
            images=sheet_images,
            output_dir=sheet_output_dir,
            model=args.model,
            base_url=args.base_url,
            limit=sheet_limit,
            sleep_seconds=args.sleep,
            max_retries=args.max_retries,
            resume=not args.no_resume,
        )
        results_by_sheet[sheet] = results
        excel_path = write_excel(results, sheet_output_dir)
        print(f"分析完成：{sheet_output_dir / 'analysis_results.jsonl'}")
        print(f"Excel 输出：{excel_path}")

    combined_excel_path = write_combined_excel(results_by_sheet, output_dir)
    print(f"多 sheet Excel 总表：{combined_excel_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
